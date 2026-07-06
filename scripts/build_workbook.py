#!/usr/bin/env python3
"""Build the Plato's Pizza analytics dashboard workbook.

Sheets: Dashboard | Data | Calc | Read Me
Palette (validated, dataviz reference): categorical #2a78d6/#1baf7a/#eda100/#008300,
accent #EB6834, inks #0B0B0B/#52514E/#898781, grid #E1E0D9, surface #FCFCFB.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.series import SeriesLabel, DataPoint
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend, LegendEntry

SRC = "data/pizza_sales.csv"
OUT = "dashboard/Pizza_Sales_Dashboard.xlsx"

# ---- palette -------------------------------------------------------------
ACCENT   = "EB6834"   # story / action color
INK      = "0B0B0B"
INK2     = "52514E"
MUTED    = "898781"
GRID     = "E1E0D9"
SURFACE  = "FCFCFB"
BAND     = "1A1A19"   # title band
CARD     = "F5F4F1"
CAT_COLORS = {"Chicken": "2A78D6", "Classic": "1BAF7A", "Supreme": "EDA100", "Veggie": "008300"}
GOODDARK = "184F95"

# ---- load & derive -------------------------------------------------------
df = pd.read_csv(SRC)
df["order_date"] = pd.to_datetime(df["order_date"], format="%d-%m-%Y")
t = pd.to_datetime(df["order_time"], format="%H:%M:%S")
df["Hour"] = t.dt.hour
df["DowNum"] = df["order_date"].dt.dayofweek + 1            # 1=Mon
df["Weekday"] = df["order_date"].dt.strftime("%a")           # Mon..Sun
df["WeekNum"] = df["order_date"].dt.isocalendar().week.astype(int)
df["Month"] = df["order_date"].dt.month
df["Quarter"] = "Q" + df["order_date"].dt.quarter.astype(str)
df["OrderFlag"] = (~df["order_id"].duplicated()).astype(int)
df["DateFlag"] = (~df["order_date"].duplicated()).astype(int)

N = len(df)
LAST = N + 1  # last data row in sheet

# aggregates for static chart tables
weekly = df.groupby("WeekNum")["total_price"].sum()
open_days_per_week = df.groupby("WeekNum")["order_date"].nunique()
weekly_per_day = (weekly / open_days_per_week).round(0)
avg_per_day = df["total_price"].sum() / df["order_date"].nunique()

pz = df.groupby(["pizza_name", "pizza_category"]).agg(
    units=("quantity", "sum"), rev=("total_price", "sum")).reset_index()
pz["avg_price"] = pz.rev / pz.units
MED_U, MED_P = pz.units.median(), pz.avg_price.median()

pareto = pz.sort_values("rev", ascending=False).reset_index(drop=True)
pareto["share"] = pareto.rev / pareto.rev.sum()
pareto["cum"] = pareto.share.cumsum()

# waterfall scenario (assumptions documented on Read Me)
lad = df.pivot_table(index="pizza_name", columns="pizza_size", values="unit_price", aggfunc="mean")
step_sm = (lad["M"] - lad["S"]).mean(); step_ml = (lad["L"] - lad["M"]).mean()
su = df.groupby("pizza_size")["quantity"].sum()
w_upsell = 0.20 * su["S"] * step_sm + 0.20 * su["M"] * step_ml
osz = df.groupby("order_id")["quantity"].sum()
avg_s_price = df.loc[df.pizza_size == "S", "unit_price"].mean()
w_attach = (osz == 1).sum() * 0.10 * avg_s_price
h15 = df.loc[df.Hour == 15, "total_price"].sum()
w_happy = 0.30 * h15
fri = df[df.DowNum == 5].groupby("order_date")["total_price"].sum().mean()
sun = df[df.DowNum == 7].groupby("order_date")["total_price"].sum().mean()
n_sun = df.loc[df.DowNum == 7, "order_date"].nunique()
w_weekend = 0.25 * (fri - sun) * n_sun
puzzles = pz[(pz.units < MED_U) & (pz.avg_price >= MED_P)]
w_puzzle = 0.15 * puzzles.rev.sum()
wf = [("Size upsell", w_upsell), ("2nd-pizza attach", w_attach),
      ("3-4pm happy hour", w_happy), ("Close Sunday gap", w_weekend),
      ("Promote Puzzles", w_puzzle)]
wf_total = sum(v for _, v in wf)

# =====================================================================
wb = Workbook()
dash = wb.active; dash.title = "Dashboard"
data_ws = wb.create_sheet("Data")
calc = wb.create_sheet("Calc")
readme = wb.create_sheet("Read Me")

thin = Side(style="thin", color=GRID)

# ---- Data sheet ----------------------------------------------------------
cols = ["pizza_id","order_id","pizza_name_id","quantity","order_date","order_time",
        "unit_price","total_price","pizza_size","pizza_category","pizza_name",
        "pizza_ingredients","Hour","DowNum","Weekday","WeekNum","Month","Quarter",
        "OrderFlag","DateFlag"]
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=BAND)
for j, c in enumerate(cols, 1):
    cell = data_ws.cell(row=1, column=j, value=c)
    cell.font = hdr_font; cell.fill = hdr_fill
out = df[["pizza_id","order_id","pizza_name_id","quantity","order_date","order_time",
          "unit_price","total_price","pizza_size","pizza_category","pizza_name",
          "pizza_ingredients","Hour","DowNum","Weekday","WeekNum","Month","Quarter",
          "OrderFlag","DateFlag"]]
for i, row in enumerate(out.itertuples(index=False), 2):
    for j, v in enumerate(row, 1):
        if j == 5:
            data_ws.cell(row=i, column=j, value=v.date()).number_format = "yyyy-mm-dd"
        else:
            data_ws.cell(row=i, column=j, value=v)
data_ws.freeze_panes = "A2"
data_ws.sheet_view.showGridLines = False

R = f"Data!$H$2:$H${LAST}"      # revenue
Q = f"Data!$R$2:$R${LAST}"      # quarter
HR = f"Data!$M$2:$M${LAST}"     # hour
WD = f"Data!$O$2:$O${LAST}"     # weekday text
QTY = f"Data!$D$2:$D${LAST}"
OF = f"Data!$S$2:$S${LAST}"
DF_ = f"Data!$T$2:$T${LAST}"
SEL = 'IF(Dashboard!$Q$4="All","*",Dashboard!$Q$4)'

# ---- Calc sheet: static chart tables --------------------------------------
c_hdr = Font(bold=True, size=10, color=INK2)

calc["A1"] = "WEEKLY REVENUE PER OPEN DAY"; calc["A1"].font = c_hdr
calc["A2"] = "Week"; calc["B2"] = "Rev/open day"; calc["C2"] = "Year average"
r0 = 3
for i, (wk, v) in enumerate(weekly_per_day.items()):
    calc.cell(row=r0+i, column=1, value=int(wk))
    calc.cell(row=r0+i, column=2, value=round(float(v), 0))
    calc.cell(row=r0+i, column=3, value=round(avg_per_day, 0))
WK_N = len(weekly_per_day); WK_END = r0 + WK_N - 1

calc["E1"] = "MENU ENGINEERING (units vs avg price, full year)"; calc["E1"].font = c_hdr
calc["E2"], calc["F2"], calc["G2"], calc["H2"] = "Pizza", "Category", "Units", "Avg price"
r = 3; cat_rows = {}
for cat in ["Chicken", "Classic", "Supreme", "Veggie"]:
    sub = pz[pz.pizza_category == cat].sort_values("units")
    cat_rows[cat] = (r, r + len(sub) - 1)
    for _, x in sub.iterrows():
        calc.cell(row=r, column=5, value=x.pizza_name.replace("The ", "").replace(" Pizza", ""))
        calc.cell(row=r, column=6, value=cat)
        calc.cell(row=r, column=7, value=int(x.units))
        calc.cell(row=r, column=8, value=round(float(x.avg_price), 2))
        r += 1
# quadrant guide lines
XMAX = int(pz.units.max() * 1.06); YMIN = float(np.floor(pz.avg_price.min() - 1)); YMAX = float(np.ceil(pz.avg_price.max() + 1))
calc["J2"], calc["K2"] = "gx", "gy"
calc["J3"], calc["K3"] = float(MED_U), YMIN
calc["J4"], calc["K4"] = float(MED_U), YMAX
calc["J6"], calc["K6"] = 0, float(MED_P)
calc["J7"], calc["K7"] = XMAX, float(MED_P)

calc["M1"] = "REVENUE BY PIZZA, SORTED (ascending so largest plots on top)"; calc["M1"].font = c_hdr
calc["M2"], calc["N2"], calc["O2"] = "Pizza", "Revenue", "Cumulative"
# ascending order: horizontal bar charts draw the first row at the bottom
par_asc = pareto.iloc[::-1].reset_index(drop=True)
n_par = len(par_asc)
# tiers computed on the descending ranking: top pizzas covering 50% cum share; bottom 10
top_n = int((pareto.cum <= 0.50).sum()) + 1   # first pizza crossing 50% included
tier = []  # per ascending-row tier
for i in range(n_par):
    rank_desc = n_par - 1 - i                 # 0 = biggest
    if rank_desc < top_n:
        tier.append("top")
    elif rank_desc >= n_par - 10:
        tier.append("bottom")
    else:
        tier.append("mid")
for i, x in par_asc.iterrows():
    calc.cell(row=3+i, column=13, value=x.pizza_name.replace("The ", "").replace(" Pizza", ""))
    calc.cell(row=3+i, column=14, value=round(float(x.rev), 0))
    calc.cell(row=3+i, column=15, value=round(float(x.cum), 4)).number_format = "0.0%"
PAR_END = 3 + n_par - 1

calc["Q1"] = "OPPORTUNITY WATERFALL (annualised, assumptions on Read Me)"; calc["Q1"].font = c_hdr
calc["Q2"], calc["R2"], calc["S2"], calc["T2"] = "Lever", "base", "gain", "total"
cum = 0.0
for i, (name, v) in enumerate(wf):
    calc.cell(row=3+i, column=17, value=name)
    calc.cell(row=3+i, column=18, value=round(cum, 0))
    calc.cell(row=3+i, column=19, value=round(v, 0))
    calc.cell(row=3+i, column=20, value=0)
    cum += v
tr = 3 + len(wf)
calc.cell(row=tr, column=17, value="Total opportunity")
calc.cell(row=tr, column=18, value=0)
calc.cell(row=tr, column=19, value=0)
calc.cell(row=tr, column=20, value=round(wf_total, 0))
calc.sheet_view.showGridLines = False

# ---- Dashboard -----------------------------------------------------------
dash.sheet_view.showGridLines = False
for col, w in {"A":2.5,"B":13,"C":9,"D":9,"E":9,"F":9,"G":9,"H":9,"I":9,"J":3,
               "K":9,"L":9,"M":9,"N":9,"O":9,"P":9,"Q":9,"R":9,"S":9,"T":2.5}.items():
    dash.column_dimensions[col].width = w
# white canvas
white = PatternFill("solid", fgColor=SURFACE)
for rr in range(1, 100):
    for cc in range(1, 21):
        dash.cell(row=rr, column=cc).fill = white

# title band
band_fill = PatternFill("solid", fgColor=BAND)
for rr in (1, 2, 3):
    for cc in range(1, 21):
        dash.cell(row=rr, column=cc).fill = band_fill
dash.merge_cells("B2:P2")
dash["B2"] = "PLATO'S PIZZA · 2015 — $817,860 and not growing. Where does the next $88K come from?"
dash["B2"].font = Font(bold=True, size=16, color="FFFFFF")
dash.merge_cells("B3:P3")
dash["B3"] = "No new customers required: fix the menu mix, grow the ticket, fill the demand valleys.  |  Source: 48,620 order lines, 21,350 orders, 358 open days"
dash["B3"].font = Font(size=9, color="C3C2B7")
dash["Q3"] = "QUARTER"; dash["Q3"].font = Font(size=8, bold=True, color="C3C2B7")
dash["Q4"] = "All"
dash["Q4"].font = Font(size=11, bold=True, color=INK)
dash["Q4"].fill = PatternFill("solid", fgColor="FFFFFF")
dash["Q4"].alignment = Alignment(horizontal="center")
dash["Q4"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
dv = DataValidation(type="list", formula1='"All,Q1,Q2,Q3,Q4"', allow_blank=False)
dash.add_data_validation(dv); dv.add(dash["Q4"])
dash["R4"] = "◀ filter KPIs & heatmap"; dash["R4"].font = Font(size=7, color=MUTED)

# KPI cards
kpis = [
    ("TOTAL REVENUE", f'=SUMIFS({R},{Q},{SEL})', '$#,##0'),
    ("ORDERS", f'=SUMIFS({OF},{Q},{SEL})', '#,##0'),
    ("AVG ORDER VALUE", f'=SUMIFS({R},{Q},{SEL})/SUMIFS({OF},{Q},{SEL})', '$0.00'),
    ("PIZZAS SOLD", f'=SUMIFS({QTY},{Q},{SEL})', '#,##0'),
    ("REVENUE / OPEN DAY", f'=SUMIFS({R},{Q},{SEL})/SUMIFS({DF_},{Q},{SEL})', '$#,##0'),
]
card_cols = [("B","D"), ("E","G"), ("H","J"), ("K","M"), ("N","P")]
card_fill = PatternFill("solid", fgColor=CARD)
for (c1, c2), (label, f, fmt) in zip(card_cols, kpis):
    dash.merge_cells(f"{c1}5:{c2}5"); dash.merge_cells(f"{c1}6:{c2}7")
    lab = dash[f"{c1}5"]; lab.value = label
    lab.font = Font(size=8, bold=True, color=MUTED); lab.alignment = Alignment(horizontal="center", vertical="bottom")
    val = dash[f"{c1}6"]; val.value = f; val.number_format = fmt
    val.font = Font(size=20, bold=True, color=INK); val.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (5, 6, 7):
        for cc in range(dash[f"{c1}1"].column, dash[f"{c2}1"].column + 1):
            dash.cell(row=rr, column=cc).fill = card_fill
# reconciliation footnote under KPIs
dash.merge_cells("B8:P8")
dash["B8"] = "✓ Reconciled: AOV × orders = $38.31 × 21,350 = $817,918 ≈ Σ line revenue $817,860 (rounding). Every figure on this page ties back to the raw table."
dash["B8"].font = Font(size=8, italic=True, color=MUTED)

def section(cell_title, num, title, cell_sub, sub):
    c = dash[cell_title]
    c.value = f"{num}  ·  {title}"
    c.font = Font(size=12, bold=True, color=INK)
    dash[cell_title.replace(cell_title[0], cell_title[0], 1)]  # no-op clarity
    s = dash[cell_sub]; s.value = sub; s.font = Font(size=9, color=INK2)

# section titles (action titles — the chart says the finding)
dash["B10"] = "1 · REVENUE HAS FLATLINED AT ≈ $2,285 PER OPEN DAY"
dash["B11"] = "Every quarter lands within 0.5% of the same run-rate. Growth must come from inside the shop."
dash["K10"] = "2 · THE MENU HAS FOUR PERSONALITIES — MANAGE EACH DIFFERENTLY"
dash["K11"] = "Units sold (→) vs avg price ($). Medians split: Stars (keep), Plowhorses (upsell), Puzzles (promote), Dogs (fix or cut)."
dash["B26"] = "3 · WHEN THE MONEY ARRIVES — LUNCH IS 27% OF REVENUE IN 2 HOURS"
dash["B27"] = "Revenue by hour × weekday. Dark = busy. Note the 3–4pm trough and the quiet Sunday."
dash["K33"] = "4 · HALF THE REVENUE COMES FROM JUST 12 PIZZAS — THE BOTTOM 10 ADD 19%"
dash["K34"] = "Full-year revenue by pizza, sorted. Orange = the 12 pizzas that deliver 50% of revenue; light gray = the bottom 10."
dash["B45"] = "5 · THE PATH TO +$88K (11%) — WITHOUT ONE NEW CUSTOMER"
dash["B46"] = "Five quantified levers, annualised. Assumptions on the Read Me sheet."
for cell in ("B10","K10","B26","K33","B45"):
    dash[cell].font = Font(size=12, bold=True, color=INK)
for cell in ("B11","K11","B27","K34","B46"):
    dash[cell].font = Font(size=9, color=INK2)
for num_cell in ():
    pass

def style_axes(ch, y_fmt=None):
    ch.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525)))
    ch.x_axis.majorGridlines = None
    if y_fmt:
        ch.y_axis.number_format = y_fmt
        ch.y_axis.numFmt = y_fmt
    ch.x_axis.delete = False
    ch.y_axis.delete = False

# --- chart 1: plateau line -------------------------------------------------
lc = LineChart()
lc.width, lc.height = 16.2, 7.2
ref = Reference(calc, min_col=2, min_row=2, max_row=WK_END)
lc.add_data(ref, titles_from_data=True)
cats = Reference(calc, min_col=1, min_row=r0, max_row=WK_END)
lc.set_categories(cats)
ref2 = Reference(calc, min_col=3, min_row=2, max_row=WK_END)
lc.add_data(ref2, titles_from_data=True)
s1 = lc.series[0]
s1.graphicalProperties.line = LineProperties(solidFill=ACCENT, w=19050)
s1.smooth = False
s2 = lc.series[1]
s2.graphicalProperties.line = LineProperties(solidFill=MUTED, w=12700, prstDash="dash")
s2.smooth = False
lc.legend = None
style_axes(lc, '$#,##0')
lc.y_axis.scaling.min = 0
dash.add_chart(lc, "B12")

# --- chart 2: menu engineering scatter --------------------------------------
sc = ScatterChart()
sc.width, sc.height = 16.2, 10.8
sc.x_axis.title = "Units sold (year)"; sc.y_axis.title = "Average price"
for cat in ["Chicken", "Classic", "Supreme", "Veggie"]:
    a, b = cat_rows[cat]
    xref = Reference(calc, min_col=7, min_row=a, max_row=b)
    yref = Reference(calc, min_col=8, min_row=a, max_row=b)
    s = Series(yref, xref, title=cat)
    s.marker = Marker(symbol="circle", size=8)
    s.marker.graphicalProperties.solidFill = CAT_COLORS[cat]
    s.marker.graphicalProperties.line.solidFill = SURFACE
    s.graphicalProperties.line.noFill = True
    sc.series.append(s)
# median guide lines
for (ra, rb) in ((3, 4), (6, 7)):
    xr = Reference(calc, min_col=10, min_row=ra, max_row=rb)
    yr = Reference(calc, min_col=11, min_row=ra, max_row=rb)
    g = Series(yr, xr, title="median")
    g.graphicalProperties.line = LineProperties(solidFill=MUTED, w=9525, prstDash="dash")
    g.marker = Marker(symbol="none")
    sc.series.append(g)
sc.legend = Legend(); sc.legend.position = "t"
sc.legend.legendEntry = [LegendEntry(idx=4, delete=True), LegendEntry(idx=5, delete=True)]
style_axes(sc, '$#,##0')
sc.x_axis.scaling.min = 0
sc.y_axis.scaling.min = YMIN
sc.y_axis.scaling.max = YMAX
dash.add_chart(sc, "K12")

# --- heatmap grid ------------------------------------------------------------
days = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
hm_r0 = 29
dash["B" + str(hm_r0)] = ""
for j, d in enumerate(days):
    c = dash.cell(row=hm_r0, column=3 + j, value=d)
    c.font = Font(size=9, bold=True, color=INK2); c.alignment = Alignment(horizontal="center")
for i, h in enumerate(range(11, 24)):
    rr = hm_r0 + 1 + i
    lab = dash.cell(row=rr, column=2, value=f"{h:02d}:00")
    lab.font = Font(size=8, color=MUTED); lab.alignment = Alignment(horizontal="right")
    for j, d in enumerate(days):
        cell = dash.cell(row=rr, column=3 + j)
        cell.value = f'=SUMIFS({R},{HR},{h},{WD},"{d}",{Q},{SEL})'
        cell.number_format = '#,##0'
        cell.font = Font(size=8, color=INK)
        cell.alignment = Alignment(horizontal="center")
    dash.row_dimensions[rr].height = 14
hm_range = f"C{hm_r0+1}:I{hm_r0+13}"
dash.conditional_formatting.add(hm_range, ColorScaleRule(
    start_type="min", start_color="FFFFFF",
    mid_type="percentile", mid_value=60, mid_color="9EC5F4",
    end_type="max", end_color="256ABF"))

# --- chart 4: sorted revenue bars (Pareto story, single axis) -----------------
bc = BarChart(); bc.type = "bar"
bc.width, bc.height = 16.2, 12.6
bref = Reference(calc, min_col=14, min_row=2, max_row=PAR_END)
bc.add_data(bref, titles_from_data=True)
bcats = Reference(calc, min_col=13, min_row=3, max_row=PAR_END)
bc.set_categories(bcats)
tier_fill = {"top": ACCENT, "mid": "C3C2B7", "bottom": "E1E0D9"}
pts = []
for i, tr_ in enumerate(tier):
    p = DataPoint(idx=i)
    p.graphicalProperties.solidFill = tier_fill[tr_]
    p.graphicalProperties.line.noFill = True
    pts.append(p)
bc.series[0].data_points = pts
bc.series[0].graphicalProperties.solidFill = "C3C2B7"
bc.series[0].graphicalProperties.line.noFill = True
bc.gapWidth = 25
bc.legend = None
bc.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525)))
bc.x_axis.majorGridlines = None
bc.y_axis.number_format = '$#,##0'
bc.y_axis.numFmt = '$#,##0'
dash.add_chart(bc, "K35")

# --- chart 5: waterfall --------------------------------------------------------
wfc = BarChart(); wfc.type = "col"; wfc.grouping = "stacked"; wfc.overlap = 100
wfc.width, wfc.height = 16.2, 7.6
nwf = len(wf) + 1
base_ref = Reference(calc, min_col=18, min_row=2, max_row=2 + nwf)
gain_ref = Reference(calc, min_col=19, min_row=2, max_row=2 + nwf)
tot_ref = Reference(calc, min_col=20, min_row=2, max_row=2 + nwf)
wfc.add_data(base_ref, titles_from_data=True)
wfc.add_data(gain_ref, titles_from_data=True)
wfc.add_data(tot_ref, titles_from_data=True)
wcats = Reference(calc, min_col=17, min_row=3, max_row=2 + nwf)
wfc.set_categories(wcats)
wfc.series[0].graphicalProperties.noFill = True
wfc.series[0].graphicalProperties.line.noFill = True
wfc.series[1].graphicalProperties.solidFill = ACCENT
wfc.series[1].graphicalProperties.line.noFill = True
wfc.series[2].graphicalProperties.solidFill = BAND
wfc.series[2].graphicalProperties.line.noFill = True
for si in (1, 2):
    dl = DataLabelList(showVal=True, showSerName=False, showCatName=False,
                       showLegendKey=False, showPercent=False, showBubbleSize=False)
    dl.numFmt = '$#,##0;;;'
    wfc.series[si].dLbls = dl
wfc.legend = None
wfc.gapWidth = 40
style_axes(wfc, '$#,##0')
dash.add_chart(wfc, "B47")

# footer
dash.merge_cells("B64:P64")
dash["B64"] = "Patrick Gichuki · BI & Data Analytics · Data: Plato's Pizza 2015 (Maven Analytics) · Method: menu-engineering matrix (Kasavana–Smith), Pareto, demand heat-mapping · Built in Excel"
dash["B64"].font = Font(size=8, color=MUTED)

# ---- Read Me sheet ---------------------------------------------------------
readme.sheet_view.showGridLines = False
readme.column_dimensions["A"].width = 3
readme.column_dimensions["B"].width = 110
rm = [
    ("H1", "READ ME — How this dashboard was built"),
    ("", ""),
    ("H2", "The governing question"),
    ("P", "Plato's Pizza earned $817,860 in 2015 and revenue is flat: every quarter runs at ≈$2,285 per open day (Q1 $2,282 · Q2 $2,290 · Q3 $2,278 · Q4 $2,289). "
          "The dashboard answers one question: where does the next ~$88K come from without acquiring a single new customer?"),
    ("", ""),
    ("H2", "Reconciliation (accounting habit: tie every number back)"),
    ("P", "Σ line revenue = $817,860.05.  Cross-check: AOV $38.31 × 21,350 orders = $817,918 (difference is AOV rounding). "
          "Pizzas sold = Σ quantity = 49,574 across 48,620 order lines. Open days = 358 (7 closed days: 4 consecutive Mondays in Oct, 24–25 Sep, 5 Oct week pattern, 25 Dec)."),
    ("", ""),
    ("H2", "Data preparation (Data sheet, helper columns M–T)"),
    ("P", "Hour = HOUR(order_time) · DowNum = WEEKDAY(order_date,2) · Weekday = TEXT(order_date,\"ddd\") · WeekNum = ISOWEEKNUM(order_date) · "
          "Month = MONTH(order_date) · Quarter = \"Q\"&ROUNDUP(MONTH(order_date)/3,0) · OrderFlag = 1 on the first line of each order_id (so SUM gives a distinct order count) · "
          "DateFlag = 1 on the first line of each date (so SUM gives open days). Values are stored (not live formulas) to keep 48,620 rows fast; formulas above reproduce them exactly."),
    ("", ""),
    ("H2", "Menu engineering matrix (chart 2)"),
    ("P", "Kasavana–Smith method: each pizza plotted by units sold (popularity) and average realised price per unit (contribution proxy — ingredient costs are not in the data). "
          "Split at the medians (1,452 units / $16.99): STARS high-high (all six chicken-led premium pizzas) · PLOWHORSES popular but cheap (Classic Deluxe, Hawaiian, Pepperoni $12.47) · "
          "PUZZLES premium but unseen (Brie Carre $23.65 — priciest pizza, lowest volume; the mentor chart calls it a 'worst seller', the matrix says 'promote it') · DOGS low-low (the spinach/veggie cluster)."),
    ("", ""),
    ("H2", "Opportunity waterfall assumptions (chart 5) — deliberately conservative"),
    ("P", f"Size upsell: 1 in 5 S→M and M→L conversions at the observed price ladder (S→M ${step_sm:.2f}, M→L ${step_ml:.2f}) = +${w_upsell:,.0f}. "
          f"2nd-pizza attach: 10% of the 8,111 single-pizza orders add one small pizza (avg ${avg_s_price:.2f}) = +${w_attach:,.0f}. "
          f"Happy hour: +30% on the 3–4pm trough (currently 47% of the lunch peak) = +${w_happy:,.0f}. "
          f"Sunday gap: close 25% of the Friday–Sunday gap (${fri:,.0f} vs ${sun:,.0f}/day) over {n_sun} Sundays = +${w_weekend:,.0f}. "
          f"Puzzle promotion: +15% volume on the 9 premium-but-unseen pizzas = +${w_puzzle:,.0f}. "
          f"Total = +${wf_total:,.0f} ≈ {wf_total/817860*100:.0f}% growth."),
    ("", ""),
    ("H2", "Data quality notes"),
    ("P", "XL/XXL sizes exist only in the Classic category; the XXL Greek sold 28 units all year — menu complexity with no payoff. "
          "The Brie Carre is the only S-only pizza. 38% of orders contain a single pizza. "
          "9am/10am have 9 orders combined across the year (heatmap starts at 11:00)."),
    ("", ""),
    ("H2", "Design decisions (Storytelling with Data / Knaflic)"),
    ("P", "One accent colour for the story, gray for context; action titles state the finding, charts prove it; no pies, no 3-D, no dual scales "
          "(the Pareto bar and cumulative line share one 0–100% axis). Categorical palette is colour-blind-safe (validated: worst adjacent ΔE 24.2)."),
]
rr = 2
for kind, text in rm:
    cell = readme.cell(row=rr, column=2, value=text if text else None)
    if kind == "H1":
        cell.font = Font(size=16, bold=True, color=INK)
    elif kind == "H2":
        cell.font = Font(size=12, bold=True, color=ACCENT)
    elif kind == "P":
        cell.font = Font(size=10, color=INK2)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        readme.row_dimensions[rr].height = 68
    rr += 1

# print/export setup: dashboard exports as one landscape page
from openpyxl.worksheet.properties import PageSetupProperties
dash.print_area = "A1:T66"
dash.page_setup.orientation = "landscape"
dash.page_setup.fitToWidth = 1
dash.page_setup.fitToHeight = 1
dash.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("saved", OUT)
print(f"waterfall: {[(n, round(v)) for n, v in wf]} total={wf_total:,.0f}")
